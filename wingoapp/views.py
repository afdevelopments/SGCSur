import json

from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout
from django.http.response import HttpResponseRedirect
from django.views import generic
from django.urls import reverse
from django.db.models import Q
from .models import *
from .forms import *
from django.views.generic import (
    ListView,
    View,
    CreateView,
    UpdateView, TemplateView
)


# dashboard pages

@login_required(login_url="/login")
def indexPage(request):
    dashboard(request)

    return redirect('/dashboard_default')


@login_required(login_url="/login")
def dashboard_default(request):
    context = {"breadcrumb": {"parent": "Dashboard", "child": "Default"}}
    return render(request, 'general/dashboard/default/index.html', context)


@login_required(login_url="/login")
def dashboard_ecommerce(request):
    context = {"breadcrumb": {"parent": "Dashboard", "child": "Ecommerce"}}
    return render(request, 'general/dashboard/ecommerce/dashboard-02.html', context)


@login_required(login_url="/login")
def dashboard_project(request):
    context = {"breadcrumb": {"parent": "Dashboard", "child": "project-Dashboard"}}
    return render(request, 'general/dashboard/project/dashboard-03.html', context)


# widgets pages


@login_required(login_url="/login")
def widgets_general(request):
    context = {"breadcrumb": {"parent": "Widgets", "child": "General"}}
    return render(request, 'widgets/general/general-widget.html', context)


@login_required(login_url="/login")
def widgets_chart(request):
    context = {"breadcrumb": {"parent": "Widgets", "child": "Chart"}}
    return render(request, 'widgets/chart/chart-widget.html', context)


# page layout views


@login_required(login_url="/login")
def page_layout_boxed(request):
    context = {"layout": "box-layout", "breadcrumb": {"parent": "Page Layout", "child": "Boxed"}}
    return render(request, 'page_layout/boxed/box-layout.html', context)


@login_required(login_url="/login")
def page_layout_rtl(request):
    context = {"layout": "rtl", "breadcrumb": {"parent": "Page Layout", "child": "RTL"}}
    return render(request, 'page_layout/RTL/layout-rtl.html', context)


@login_required(login_url="/login")
def page_layout_dark(request):
    context = {"layout": "dark-only", "breadcrumb": {"parent": "Page Layout", "child": "Layout Dark"}}
    return render(request, 'page_layout/dark_layout/layout-dark.html', context)


@login_required(login_url="/login")
def page_layout_hide_nav_scroll(request):
    context = {"breadcrumb": {"parent": "Page Layout", "child": "Hide Menu On Scroll"}}
    return render(request, 'page_layout/hide_nav_scroll/hide-on-scroll.html', context)


@login_required(login_url="/login")
def page_layout_footer_light(request):
    context = {"breadcrumb": {"parent": "Page Layout", "child": "Footer Light"}}
    return render(request, 'page_layout/footer_light/footer-light.html', context)


@login_required(login_url="/login")
def page_layout_footer_dark(request):
    context = {"footer": "footer-dark", "breadcrumb": {"parent": "Page Layout", "child": "Footer Dark"}}
    return render(request, 'page_layout/footer_dark/footer-dark.html', context)


@login_required(login_url="/login")
def page_layout_footer_fixed(request):
    context = {"footer": "footer-fix", "breadcrumb": {"parent": "Page Layout", "child": "Footer Fixed"}}
    return render(request, 'page_layout/footer_fixed/footer-fixed.html', context)


# project views


@login_required(login_url="/login")
def project_project_list(request):
    context = {"breadcrumb": {"parent": "Project", "child": "Project List"}}
    return render(request, 'project/project_list/projects.html', context)


@login_required(login_url="/login")
def project_create_new(request):
    context = {"breadcrumb": {"parent": "Project", "child": "Create New"}}
    return render(request, 'project/create_new/projectcreate.html', context)


# file manager view


@login_required(login_url="/login")
def file_manager(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "File Manager"}}
    return render(request, 'file_manager/file-manager.html', context)


# kanban board view


@login_required(login_url="/login")
def kanban_board(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "Kanban Board"}}
    return render(request, 'kanban_board/kanban.html', context)


# Ecommerce views


@login_required(login_url="/login")
def ecommerce_product_default(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Product"}}
    return render(request, 'ecommerce/product/product.html', context)


@login_required(login_url="/login")
def ecommerce_product_page(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Product Page"}}
    return render(request, 'ecommerce/product_page/product-page.html', context)


@login_required(login_url="/login")
def ecommerce_product_list(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Product List"}}
    return render(request, 'ecommerce/product_list/list-products.html', context)


@login_required(login_url="/login")
def ecommerce_payment_details(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Payment Details"}}
    return render(request, 'ecommerce/payment_details/payment-details.html', context)


@login_required(login_url="/login")
def ecommerce_order_history(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Order History"}}
    return render(request, 'ecommerce/order_history/order-history.html', context)


@login_required(login_url="/login")
def ecommerce_invoice(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Invoice"}}
    return render(request, 'ecommerce/invoice/invoice-template.html', context)


@login_required(login_url="/login")
def ecommerce_cart(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Cart"}}
    return render(request, 'ecommerce/cart/cart.html', context)


@login_required(login_url="/login")
def ecommerce_wishlist(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Wishlist"}}
    return render(request, 'ecommerce/wishlist/list-wish.html', context)


@login_required(login_url="/login")
def ecommerce_checkout(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Checkout"}}
    return render(request, 'ecommerce/checkout/checkout.html', context)


@login_required(login_url="/login")
def ecommerce_pricing(request):
    context = {"breadcrumb": {"parent": "E-Commerce", "child": "Pricing"}}
    return render(request, 'ecommerce/pricing/pricing.html', context)


# email views


@login_required(login_url="/login")
def email_email_app(request):
    context = {"breadcrumb": {"parent": "Email", "child": "Mail Inbox"}}
    return render(request, 'email/email_app/email_inbox.html', context)


@login_required(login_url="/login")
def email_read_mail(request):
    context = {"breadcrumb": {"parent": "Email", "child": "Read Mail"}}
    return render(request, 'email/read_mail/email_read.html', context)


@login_required(login_url="/login")
def email_email_compose(request):
    context = {"breadcrumb": {"parent": "Email", "child": "Email Compose"}}
    return render(request, 'email/email_compose/email_compose.html', context)


# chat views


@login_required(login_url="/login")
def chat_chat_app(request):
    context = {"breadcrumb": {"parent": "Chat", "child": "Chat App"}}
    return render(request, 'chat/chat_app/chat.html', context)


@login_required(login_url="/login")
def chat_video_chat(request):
    context = {"breadcrumb": {"parent": "Chat", "child": "Video Chat"}}
    return render(request, 'chat/video_chat/chat-video.html', context)


# users views


@login_required(login_url="/login")
def users_users_profile(request):
    context = {"breadcrumb": {"parent": "Users", "child": "User Profile"}}
    return render(request, 'users/users_profile/user-profile.html', context)


@login_required(login_url="/login")
def users_users_edit(request):
    context = {"breadcrumb": {"parent": "Users", "child": "User Edit"}}
    return render(request, 'users/users_edit/edit-profile.html', context)


@login_required(login_url="/login")
def users_users_card(request):
    context = {"breadcrumb": {"parent": "Users", "child": "User Cards"}}
    return render(request, 'users/users_cards/user-cards.html', context)


# bookmarks views


@login_required(login_url="/login")
def bookmarks(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "Bookmarks"}}
    return render(request, 'bookmarks/bookmark.html', context)


# contacts views


@login_required(login_url="/login")
def contacts(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "Contact"}}
    return render(request, 'contacts/contacts.html', context)


# task views


@login_required(login_url="/login")
def tasks(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "Task"}}
    return render(request, 'tasks/task.html', context)


# calendar views


@login_required(login_url="/login")
def calendar(request):
    context = {"breadcrumb": {"parent": "Calendar", "child": "Calendar Basic"}}
    return render(request, 'calendar/calendar-basic.html', context)


# social app views


@login_required(login_url="/login")
def social_app(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "Social App"}}
    return render(request, 'social_app/social-app.html', context)


# to do views


@login_required(login_url="/login")
def to_do_design(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "To Do"}}
    return render(request, 'to_do/to-do.html', context)


@login_required(login_url="/login")



@login_required(login_url="/login")
def markAllComplete(request):
    allTasks = Task.objects.all()
    for oneTask in allTasks:
        oneTask.complete = True
        oneTask.save()
    return HttpResponseRedirect("/to_do_database")


@login_required(login_url="/login")
def markAllIncomplete(request):
    allTasks = Task.objects.all()
    for oneTask in allTasks:
        oneTask.complete = False
        oneTask.save()
    return HttpResponseRedirect("/to_do_database")


@login_required(login_url="/login")
def deleteTask(request, pk):
    item = Task.objects.get(id=pk)

    # if request.method == "POST":
    item.delete()
    return HttpResponseRedirect("/to_do_database")


@login_required(login_url="/login")
def updateTask(request, pk):
    task = Task.objects.get(id=pk)
    if task.complete == False:
        task.complete = True
        task.save()
    else:
        task.complete = False
        task.save()

    return HttpResponseRedirect("/to_do_database")


# search website views


@login_required(login_url="/login")
def search_website(request):
    context = {"breadcrumb": {"parent": "Search Pages", "child": "Search Website"}}
    return render(request, 'search_website/search.html', context)


# forms views


@login_required(login_url="/login")
def form_form_validation(request):
    context = {"breadcrumb": {"parent": "Form Controls", "child": "Form validation"}}
    return render(request, 'forms/form_controls/form_validation/form-validation.html', context)


@login_required(login_url="/login")
def form_base_inputs(request):
    context = {"breadcrumb": {"parent": "Form Controls", "child": "Base Inputs"}}
    return render(request, 'forms/form_controls/base_inputs/base-input.html', context)


@login_required(login_url="/login")
def checkbox_and_radio(request):
    context = {"breadcrumb": {"parent": "Form Controls", "child": "Checkbox & Radio"}}
    return render(request, 'forms/form_controls/checkbox_and_radio/radio-checkbox-control.html', context)


@login_required(login_url="/login")
def input_groups(request):
    context = {"breadcrumb": {"parent": "Form Controls", "child": "Input Groups"}}
    return render(request, 'forms/form_controls/input_groups/input-group.html', context)


@login_required(login_url="/login")
def mega_options(request):
    context = {"breadcrumb": {"parent": "Form Controls", "child": "Mega Options"}}
    return render(request, 'forms/form_controls/mega_options/megaoptions.html', context)


@login_required(login_url="/login")
def datepicker(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Datepicker"}}
    return render(request, 'forms/form_widgets/datepicker/datepicker.html', context)


@login_required(login_url="/login")
def timepicker(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Timepicker"}}
    return render(request, 'forms/form_widgets/timepicker/time-picker.html', context)


@login_required(login_url="/login")
def datetimepicker(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Date & Time"}}
    return render(request, 'forms/form_widgets/dateTimePicker/datetimepicker.html', context)


@login_required(login_url="/login")
def daterangepicker(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Date Range"}}
    return render(request, 'forms/form_widgets/dateRangePicker/daterangepicker.html', context)


@login_required(login_url="/login")
def touchspin(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Touchspin"}}
    return render(request, 'forms/form_widgets/touchspin/touchspin.html', context)


@login_required(login_url="/login")
def select2(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Select2"}}
    return render(request, 'forms/form_widgets/select2/select2.html', context)


@login_required(login_url="/login")
def switch(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Switch"}}
    return render(request, 'forms/form_widgets/switch/switch.html', context)


@login_required(login_url="/login")
def typeahead(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Typehead"}}
    return render(request, 'forms/form_widgets/typeahead/typeahead.html', context)


@login_required(login_url="/login")
def clipboard(request):
    context = {"breadcrumb": {"parent": "Form Widgets", "child": "Clipboard"}}
    return render(request, 'forms/form_widgets/clipboard/clipboard.html', context)


@login_required(login_url="/login")
def default_form(request):
    context = {"breadcrumb": {"parent": "Form Layout", "child": "Default Forms"}}
    return render(request, 'forms/form_layout/default_forms/default-form.html', context)


@login_required(login_url="/login")
def form_wizard_1(request):
    context = {"breadcrumb": {"parent": "Form Layout", "child": "Form Wizard"}}
    return render(request, 'forms/form_layout/form_wizard_1/form-wizard.html', context)


@login_required(login_url="/login")
def form_wizard_2(request):
    context = {"breadcrumb": {"parent": "Form Layout", "child": "Form Wizard"}}
    return render(request, 'forms/form_layout/form_wizard_2/form-wizard-two.html', context)


@login_required(login_url="/login")
def form_wizard_3(request):
    context = {"breadcrumb": {"parent": "Form Layout", "child": "Form Wizard"}}
    return render(request, 'forms/form_layout/form_wizard_3/form-wizard-three.html', context)


# tables views

# bootstrap tables


@login_required(login_url="/login")
def bootstrap_basic_tables(request):
    context = {"breadcrumb": {"parent": "Bootstrap Tables", "child": "Basic Tables"}}
    return render(request, 'tables/bootstrap_tables/basic_tables/bootstrap-basic-table.html', context)


@login_required(login_url="/login")
def bootstrap_border_tables(request):
    context = {"breadcrumb": {"parent": "Bootstrap Tables", "child": "Border tables"}}
    return render(request, 'tables/bootstrap_tables/border_tables/bootstrap-border-table.html', context)


@login_required(login_url="/login")
def bootstrap_sizing_tables(request):
    context = {"breadcrumb": {"parent": "Bootstrap Tables", "child": "Sizing Tables"}}
    return render(request, 'tables/bootstrap_tables/sizing_tables/bootstrap-sizing-table.html', context)


@login_required(login_url="/login")
def bootstrap_styling_tables(request):
    context = {"breadcrumb": {"parent": "Bootstrap Tables", "child": "Styling Tables"}}
    return render(request, 'tables/bootstrap_tables/styling_tables/bootstrap-styling-table.html', context)


@login_required(login_url="/login")
def bootstrap_table_components(request):
    context = {"breadcrumb": {"parent": "Bootstrap Tables", "child": "Table Components"}}
    return render(request, 'tables/bootstrap_tables/table_components/table-components.html', context)


# data tables


@login_required(login_url="/login")
def data_tables_advance_init(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "Advance Data Tables"}}
    return render(request, 'tables/data_tables/advance_init/datatable-advance.html', context)


@login_required(login_url="/login")
def data_tables_AJAX(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "Ajax Data Tables"}}
    return render(request, 'tables/data_tables/AJAX/datatable-AJAX.html', context)


@login_required(login_url="/login")
def data_tables_API(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "API Data Tables"}}
    return render(request, 'tables/data_tables/API/datatable-API.html', context)


@login_required(login_url="/login")
def data_tables_basic_init(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "Basic DataTables"}}
    return render(request, 'tables/data_tables/basic_init/datatable-basic-init.html', context)


@login_required(login_url="/login")
def data_tables_data_source(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "DATA Source DataTables"}}
    return render(request, 'tables/data_tables/data_source/datatable-data-source.html', context)


@login_required(login_url="/login")
def data_tables_plug_in(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "Plugin DataTable"}}
    return render(request, 'tables/data_tables/plug_in/datatable-plugin.html', context)


@login_required(login_url="/login")
def data_tables_server_side(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "Server Side DataTable"}}
    return render(request, 'tables/data_tables/server_side/datatable-server-side.html', context)


@login_required(login_url="/login")
def data_tables_styling(request):
    context = {"breadcrumb": {"parent": "Data Tables", "child": "Styling DataTables"}}
    return render(request, 'tables/data_tables/styling/datatable-styling.html', context)


# ex datatables


@login_required(login_url="/login")
def ex_data_tables_auto_fill(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Autofill Datatables"}}
    return render(request, 'tables/ex_data_tables/auto_fill/datatable-ext-autofill.html', context)


@login_required(login_url="/login")
def ex_data_tables_basic_button(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Basic Button"}}
    return render(request, 'tables/ex_data_tables/basic_button/datatable-ext-basic-button.html', context)


@login_required(login_url="/login")
def ex_data_tables_column_reorder(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Column Reorder"}}
    return render(request, 'tables/ex_data_tables/column_reorder/datatable-ext-col-reorder.html', context)


@login_required(login_url="/login")
def ex_data_tables_fixed_header(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Fixed Header"}}
    return render(request, 'tables/ex_data_tables/fixed_header/datatable-ext-fixed-header.html', context)


@login_required(login_url="/login")
def ex_data_tables_html_5_export(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "HTML 5 Data Export"}}
    return render(request, 'tables/ex_data_tables/html_5_export/datatable-ext-html-5-data-export.html', context)


@login_required(login_url="/login")
def ex_data_tables_key_table(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Key Table"}}
    return render(request, 'tables/ex_data_tables/key_table/datatable-ext-key-table.html', context)


@login_required(login_url="/login")
def ex_data_tables_responsive(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Responsive Datatables"}}
    return render(request, 'tables/ex_data_tables/responsive/datatable-ext-responsive.html', context)


@login_required(login_url="/login")
def ex_data_tables_row_reorder(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Row Reorder"}}
    return render(request, 'tables/ex_data_tables/row_reorder/datatable-ext-row-reorder.html', context)


@login_required(login_url="/login")
def ex_data_tables_scroller(request):
    context = {"breadcrumb": {"parent": "Extension Data Tables", "child": "Scroller"}}
    return render(request, 'tables/ex_data_tables/scroller/datatable-ext-scroller.html', context)


# js grid table


@login_required(login_url="/login")
def js_grid_table(request):
    context = {"breadcrumb": {"parent": "Tables", "child": "JS Grid Tables"}}
    return render(request, 'tables/js_grid_table/jsgrid-table.html', context)


# ui kits views


@login_required(login_url="/login")
def ui_kits_accordion(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Accordion"}}
    return render(request, 'ui_kits/according.html', context)


@login_required(login_url="/login")
def ui_kits_alert(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Alert"}}
    return render(request, 'ui_kits/alert.html', context)


@login_required(login_url="/login")
def ui_kits_avatars(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Avatars"}}
    return render(request, 'ui_kits/avatars.html', context)


@login_required(login_url="/login")
def ui_kits_dropdown(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Dropdown"}}
    return render(request, 'ui_kits/dropdown.html', context)


@login_required(login_url="/login")
def ui_kits_grid(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Grid"}}
    return render(request, 'ui_kits/grid.html', context)


@login_required(login_url="/login")
def ui_kits_helper_classes(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Helper Classes"}}
    return render(request, 'ui_kits/helper-classes.html', context)


@login_required(login_url="/login")
def ui_kits_list(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "List"}}
    return render(request, 'ui_kits/list.html', context)


@login_required(login_url="/login")
def ui_kits_loader(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Loader"}}
    return render(request, 'ui_kits/loader.html', context)


@login_required(login_url="/login")
def ui_kits_modal(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Modal"}}
    return render(request, 'ui_kits/modal.html', context)


@login_required(login_url="/login")
def ui_kits_popover(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Popover"}}
    return render(request, 'ui_kits/popover.html', context)


@login_required(login_url="/login")
def ui_kits_progress_bar(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Progress Bar"}}
    return render(request, 'ui_kits/progress-bar.html', context)


@login_required(login_url="/login")
def ui_kits_state_color(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "State Color"}}
    return render(request, 'ui_kits/state-color.html', context)


@login_required(login_url="/login")
def ui_kits_tag_pills(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Tag Pills"}}
    return render(request, 'ui_kits/tag-pills.html', context)


@login_required(login_url="/login")
def ui_kits_box_shadow(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Shadow"}}
    return render(request, 'ui_kits/box-shadow.html', context)


@login_required(login_url="/login")
def ui_kits_tooltip(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Tooltip"}}
    return render(request, 'ui_kits/tooltip.html', context)


@login_required(login_url="/login")
def ui_kits_typography(request):
    context = {"breadcrumb": {"parent": "Ui Kits", "child": "Typography"}}
    return render(request, 'ui_kits/typography.html', context)


# ui kits tabs views


@login_required(login_url="/login")
def ui_kits_tabs_bootstrap(request):
    context = {"breadcrumb": {"parent": "Tabs", "child": "Bootstrap Tabs"}}
    return render(request, 'ui_kits//tabs/tab-bootstrap.html', context)


@login_required(login_url="/login")
def ui_kits_tabs_line(request):
    context = {"breadcrumb": {"parent": "Tabs", "child": "Line"}}
    return render(request, 'ui_kits//tabs/tab-material.html', context)


# bonus ui views


@login_required(login_url="/login")
def bonus_ui_basic_card(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Basic Card"}}
    return render(request, 'bonus_ui/basic-card/basic-card.html', context)


@login_required(login_url="/login")
def bonus_ui_bootstrap_notify(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Bootstrap Notify"}}
    return render(request, 'bonus_ui/bootstrap-notify/bootstrap-notify.html', context)


@login_required(login_url="/login")
def bonus_ui_breadcrumb(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Breadcrumb"}}
    return render(request, 'bonus_ui/breadcrumb/breadcrumb.html', context)


@login_required(login_url="/login")
def bonus_ui_creative_card(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Creative card"}}
    return render(request, 'bonus_ui/creative-card/creative-card.html', context)


@login_required(login_url="/login")
def bonus_ui_dragable_card(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Dragable Card"}}
    return render(request, 'bonus_ui/dragable-card/dragable-card.html', context)


@login_required(login_url="/login")
def bonus_ui_dropzone(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Dropzone"}}
    return render(request, 'bonus_ui/dropzone/dropzone.html', context)


@login_required(login_url="/login")
def bonus_ui_image_cropper(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Image Cropper"}}
    return render(request, 'bonus_ui/image-cropper/image-cropper.html', context)


@login_required(login_url="/login")
def bonus_ui_modal_animated(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Modal Animated"}}
    return render(request, 'bonus_ui/modal-animated/modal-animated.html', context)


@login_required(login_url="/login")
def bonus_ui_owl_carousel(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Owl Corousel"}}
    return render(request, 'bonus_ui/owl-carousel/owl-carousel.html', context)


@login_required(login_url="/login")
def bonus_ui_pagination(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Pagination"}}
    return render(request, 'bonus_ui/pagination/pagination.html', context)


@login_required(login_url="/login")
def bonus_ui_range_slider(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Range Slider"}}
    return render(request, 'bonus_ui/range-slider/range-slider.html', context)


@login_required(login_url="/login")
def bonus_ui_rating(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Rating"}}
    return render(request, 'bonus_ui/rating/rating.html', context)


@login_required(login_url="/login")
def bonus_ui_ribbons(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Ribbons"}}
    return render(request, 'bonus_ui/ribbons/ribbons.html', context)


@login_required(login_url="/login")
def bonus_ui_scrollable(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Scrollable"}}
    return render(request, 'bonus_ui/scrollable/scrollable.html', context)


@login_required(login_url="/login")
def bonus_ui_steps(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Steps"}}
    return render(request, 'bonus_ui/steps/steps.html', context)


@login_required(login_url="/login")
def bonus_ui_sticky(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Sticky"}}
    return render(request, 'bonus_ui/sticky/sticky.html', context)


@login_required(login_url="/login")
def bonus_ui_sweet_alert2(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Sweet Alert 2"}}
    return render(request, 'bonus_ui/sweet-alert2/sweet-alert2.html', context)


@login_required(login_url="/login")
def bonus_ui_tabbed_card(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Tabbed Card"}}
    return render(request, 'bonus_ui/tabbed-card/tabbed-card.html', context)


@login_required(login_url="/login")
def bonus_ui_tour(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Tour"}}
    return render(request, 'bonus_ui/tour/tour.html', context)


@login_required(login_url="/login")
def bonus_ui_tree(request):
    context = {"breadcrumb": {"parent": "Bonus Ui", "child": "Tree"}}
    return render(request, 'bonus_ui/tree/tree.html', context)

    # bonus ui timeline


@login_required(login_url="/login")
def bonus_ui_timeline_1(request):
    context = {"breadcrumb": {"parent": "Timeline", "child": "Timeline 1"}}
    return render(request, 'bonus_ui/timeline/timeline-v-1.html', context)


@login_required(login_url="/login")
def bonus_ui_timeline_2(request):
    context = {"breadcrumb": {"parent": "Timeline", "child": "Timeline 2"}}
    return render(request, 'bonus_ui/timeline/timeline-v-2.html', context)


# builders  views


@login_required(login_url="/login")
def button_builder(request):
    context = {"breadcrumb": {"parent": "Builders", "child": "Button Builder"}}
    return render(request, 'builders/button-builder/button-builder.html', context)


@login_required(login_url="/login")
def form_builder_1(request):
    context = {"breadcrumb": {"parent": "Builders", "child": "Form Builder 1"}}
    return render(request, 'builders/form-builder-1/form-builder-1.html', context)


@login_required(login_url="/login")
def form_builder_2(request):
    context = {"breadcrumb": {"parent": "Builders", "child": "Form Builder 2"}}
    return render(request, 'builders/form-builder-2/form-builder-2.html', context)


@login_required(login_url="/login")
def page_builder(request):
    context = {"breadcrumb": {"parent": "Builders", "child": "Page Builder"}}
    return render(request, 'builders/pagebuild/pagebuild.html', context)


# animation views


@login_required(login_url="/login")
def animate(request):
    context = {"breadcrumb": {"parent": "Animation", "child": "Animate"}}
    return render(request, 'animation/animate/animate.html', context)


@login_required(login_url="/login")
def scroll_reveal(request):
    context = {"breadcrumb": {"parent": "Animation", "child": "Scroll Reveal"}}
    return render(request, 'animation/scroll-reval/scroll-reval.html', context)


@login_required(login_url="/login")
def AOS_animation(request):
    context = {"breadcrumb": {"parent": "Animation", "child": "AOS Animation"}}
    return render(request, 'animation/AOS/AOS.html', context)


@login_required(login_url="/login")
def tilt_animation(request):
    context = {"breadcrumb": {"parent": "Animation", "child": "Tilt Animation"}}
    return render(request, 'animation/tilt/tilt.html', context)


@login_required(login_url="/login")
def wow_animation(request):
    context = {"breadcrumb": {"parent": "Animation", "child": "Wow Animation"}}
    return render(request, 'animation/wow/wow.html', context)


# icons views


@login_required(login_url="/login")
def flag_icon(request):
    context = {"breadcrumb": {"parent": "Icons", "child": "Flag Icons"}}
    return render(request, 'icons/flag-icon/flag-icon.html', context)


@login_required(login_url="/login")
def fontawesome_icon(request):
    context = {"breadcrumb": {"parent": "Icons", "child": "Fontawesome Icons"}}
    return render(request, 'icons/font-awesome/font-awesome.html', context)


@login_required(login_url="/login")
def ico_icon(request):
    context = {"breadcrumb": {"parent": "Icons", "child": "ICO Icons"}}
    return render(request, 'icons/ico-icon/ico-icon.html', context)


@login_required(login_url="/login")
def thimify_icon(request):
    context = {"breadcrumb": {"parent": "Icons", "child": "Themify Icons"}}
    return render(request, 'icons/themify-icon/themify-icon.html', context)


@login_required(login_url="/login")
def feather_icon(request):
    context = {"breadcrumb": {"parent": "Icons", "child": "Feather Icons"}}
    return render(request, 'icons/feather-icon/feather-icon.html', context)


@login_required(login_url="/login")
def whether_icon(request):
    context = {"breadcrumb": {"parent": "Icons", "child": "Weather Icon"}}
    return render(request, 'icons/whether-icon/whether-icon.html', context)


# buttons views


@login_required(login_url="/login")
def buttons(request):
    context = {"breadcrumb": {"parent": "Buttons", "child": "Default Style"}}
    return render(request, 'buttons/buttons/buttons.html', context)


@login_required(login_url="/login")
def buttons_flat(request):
    context = {"breadcrumb": {"parent": "Buttons", "child": "Button-flat"}}
    return render(request, 'buttons/buttons-flat/buttons-flat.html', context)


@login_required(login_url="/login")
def buttons_edge(request):
    context = {"breadcrumb": {"parent": "Buttons", "child": "Button-Edge"}}
    return render(request, 'buttons/buttons-edge/buttons-edge.html', context)


@login_required(login_url="/login")
def raised_button(request):
    context = {"breadcrumb": {"parent": "Buttons", "child": "Raised-Buttons"}}
    return render(request, 'buttons/raised-button/raised-button.html', context)


@login_required(login_url="/login")
def button_group(request):
    context = {"breadcrumb": {"parent": "Buttons", "child": "Button-Group"}}
    return render(request, 'buttons/button-group/button-group.html', context)


# charts views


@login_required(login_url="/login")
def apex_chart(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Apex Chart"}}
    return render(request, 'charts/chart-apex/chart-apex.html', context)


@login_required(login_url="/login")
def google_chart(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Google Chart"}}
    return render(request, 'charts/chart-google/chart-google.html', context)


@login_required(login_url="/login")
def sparkline_chart(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Spakline Chart"}}
    return render(request, 'charts/chart-sparkline/chart-sparkline.html', context)


@login_required(login_url="/login")
def flot_chart(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Flot Chart"}}
    return render(request, 'charts/chart-flot/chart-flot.html', context)


@login_required(login_url="/login")
def knob_chart(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Know Chart"}}
    return render(request, 'charts/chart-knob/chart-knob.html', context)


@login_required(login_url="/login")
def morris_chart(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Morris Chart"}}
    return render(request, 'charts/chart-morris/chart-morris.html', context)


@login_required(login_url="/login")
def chartjs(request):
    context = {"breadcrumb": {"parent": "charts", "child": "ChartJS Chart"}}
    return render(request, 'charts/chartjs/chartjs.html', context)


@login_required(login_url="/login")
def chartist(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Chartist Chart"}}
    return render(request, 'charts/chartist/chartist.html', context)


@login_required(login_url="/login")
def peity_chart(request):
    context = {"breadcrumb": {"parent": "charts", "child": "Peity Chart"}}
    return render(request, 'charts/chart-peity/chart-peity.html', context)


# landing page view


@login_required(login_url="/login")
def landing_page(request):
    return render(request, 'landing_page/landing-page.html')


# sample page view


@login_required(login_url="/login")
def sample_page(request):
    context = {"breadcrumb": {"parent": "Pages", "child": "Sample Page"}}
    return render(request, 'sample_page/sample-page.html', context)


# internationalization  view


@login_required(login_url="/login")
def internationalization(request):
    context = {"breadcrumb": {"parent": "Pages", "child": "Internationalization"}}
    return render(request, 'internationalization/internationalization.html', context)


# starter kit  view


@login_required(login_url="/login")
def starter_kit(request):
    context = {"breadcrumb": {"parent": "Color Version", "child": "Layout Light"}}
    return render(request, 'starter_kit/index.html', context)


@login_required(login_url="/login")
def starter_layout_dark(request):
    context = {"breadcrumb": {"parent": "Color Version", "child": "Layout Dark"}, "layout": "dark-only"}
    return render(request, 'starter_kit/layout-dark.html', context)


@login_required(login_url="/login")
def starter_box_layout(request):
    context = {"breadcrumb": {"parent": "Page Layout", "child": "Boxed"}, "layout": "box-layout"}
    return render(request, 'starter_kit/boxed.html', context)


@login_required(login_url="/login")
def starter_RTL(request):
    context = {"breadcrumb": {"parent": "Page Layout", "child": "RTL"}, "layout": "rtl"}
    return render(request, 'starter_kit/layout-rtl.html', context)


@login_required(login_url="/login")
def starter_footer_light(request):
    context = {"breadcrumb": {"parent": "Footers", "child": "Footer Light"}}
    return render(request, 'starter_kit/footer-light.html', context)


@login_required(login_url="/login")
def starter_footer_dark(request):
    context = {"breadcrumb": {"parent": "Footers", "child": "Footer Dark"}, "footer": "footer-dark"}
    return render(request, 'starter_kit/footer-dark.html', context)


@login_required(login_url="/login")
def starter_footer_fixed(request):
    context = {"breadcrumb": {"parent": "Footers", "child": "Footer fixed"}, "footer": "footer-fix"}
    return render(request, 'starter_kit/footer-fixed.html', context)


# others view

# error pages


@login_required(login_url="/login")
def error_page_1(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/error_page/error-page1/error-page1.html', context)


@login_required(login_url="/login")
def error_page_2(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/error_page/error-page2/error-page2.html', context)


@login_required(login_url="/login")
def error_page_3(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/error_page/error-page3/error-page3.html', context)


@login_required(login_url="/login")
def error_page_4(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/error_page/error-page4/error-page4.html', context)

    # authentication


def login_simple(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if 'next' in request.GET:
                nextPage = request.GET['next']
                return HttpResponseRedirect(nextPage)
            else:
                return redirect('dashboard_ecommerce')
    else:
        form = CustomAuthenticationForm()
    context = {"breadcrumb": {"parent": "parent", "child": "child"}, "form": form}
    return render(request, 'others/authentication/login/login.html', context)


def logout_view(request):
    logout(request)
    return redirect('login')


def login_with_bg_image(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/login-one/login_one.html', context)


def login_with_image_two(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/login-two/login_two.html', context)


def login_with_validation(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/login-bs-validation/login-bs-validation.html', context)


def login_with_tooltip(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/login-bs-tt-validation/login-bs-tt-validation.html', context)


def login_with_sweetalert(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/login-sa-validation/login-sa-validation.html', context)


def register_simple(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard_ecommerce')
    else:
        form = UserCreationForm()

    return render(request, 'others/authentication/sign-up/sign-up.html', {"form": form})


def register_with_bg_image(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/sign-up-one/sign-up-one.html', context)


def register_with_bg_video(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/sign-up-two/sign-up-two.html', context)


def unlock_user(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/unlock/unlock.html', context)


def forget_password(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/forget-password/forget-password.html', context)


def create_password(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/create-password/creat-password.html', context)


def maintenance(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/authentication/maintenance/maintenance.html', context)

    # coming soon views


@login_required(login_url="/login")
def coming_simple(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/coming_soon/comingsoon/comingsoon.html', context)


@login_required(login_url="/login")
def coming_with_bg_image(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/coming_soon/comingsoon-bg-img/comingsoon-bg-img.html', context)


@login_required(login_url="/login")
def coming_with_bg_video(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/coming_soon/comingsoon-bg-video/comingsoon-bg-video.html', context)

    # email templates views


@login_required(login_url="/login")
def basic_email(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/email_templates/basic-template.html', context)


@login_required(login_url="/login")
def basic_with_header(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/email_templates/email-header.html', context)


@login_required(login_url="/login")
def email_template(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/email_templates/template-email.html', context)


@login_required(login_url="/login")
def email_template_2(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/email_templates/template-email-2.html', context)


@login_required(login_url="/login")
def ecommerce_email(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/email_templates/ecommerce-templates.html', context)


@login_required(login_url="/login")
def order_success(request):
    context = {"breadcrumb": {"parent": "parent", "child": "child"}}
    return render(request, 'others/email_templates/email-order-success.html', context)


# gallery views


@login_required(login_url="/login")
def gallery_grid(request):
    context = {"breadcrumb": {"parent": "Gallery", "child": "Gallery Grid"}}
    return render(request, 'gallery/gallery_grid/gallery.html', context)


@login_required(login_url="/login")
def gallery_grid_desc(request):
    context = {"breadcrumb": {"parent": "Gallery", "child": "Gallery Grid With Description"}}
    return render(request, 'gallery/gallery_grid_desc/gallery-with-description.html', context)


@login_required(login_url="/login")
def masonry_gallery(request):
    context = {"breadcrumb": {"parent": "Gallery", "child": "Mansory Gallery"}}
    return render(request, 'gallery/masonry_gallery/gallery-masonry.html', context)


@login_required(login_url="/login")
def masonry_with_desc(request):
    context = {"breadcrumb": {"parent": "Gallery", "child": "Mansory Gallery With Description"}}
    return render(request, 'gallery/masonry_with_desc/masonry-gallery-with-disc.html', context)


@login_required(login_url="/login")
def hover_effect(request):
    context = {"breadcrumb": {"parent": "Gallery", "child": "Hover Effect"}}
    return render(request, 'gallery/hover_effects/gallery-hover.html', context)


# blog pages views


@login_required(login_url="/login")
def blog_details(request):
    context = {"breadcrumb": {"parent": "Blog", "child": "Blog Details"}}
    return render(request, 'blog/blog_details/blog.html', context)


@login_required(login_url="/login")
def blog_single(request):
    context = {"breadcrumb": {"parent": "Blog", "child": "Blog Single"}}
    return render(request, 'blog/blog_single/blog-single.html', context)


@login_required(login_url="/login")
def add_post(request):
    context = {"breadcrumb": {"parent": "Blog", "child": "Add Post"}}
    return render(request, 'blog/add_post/add-post.html', context)


# FAQ views


@login_required(login_url="/login")
def faq(request):
    context = {"breadcrumb": {"parent": "FAQ", "child": "FAQ"}}
    return render(request, 'faq/faq.html', context)


# job search


@login_required(login_url="/login")
def cards_view(request):
    context = {"breadcrumb": {"parent": "Job Search", "child": "Cards View"}}
    return render(request, 'job_search/cards_view/job-cards-view.html', context)


@login_required(login_url="/login")
def list_view(request):
    context = {"breadcrumb": {"parent": "Job Search", "child": "List View"}}
    return render(request, 'job_search/list_view/job-list-view.html', context)


@login_required(login_url="/login")
def job_details(request):
    context = {"breadcrumb": {"parent": "Job Search", "child": "Job Details"}}
    return render(request, 'job_search/job_details/job-details.html', context)


@login_required(login_url="/login")
def apply(request):
    context = {"breadcrumb": {"parent": "Job Search", "child": "Apply"}}
    return render(request, 'job_search/apply/job-apply.html', context)


# learning views


@login_required(login_url="/login")
def learning_list(request):
    context = {"breadcrumb": {"parent": "Learning", "child": "Learning List"}}
    return render(request, 'learning/learning_list/learning-list-view.html', context)


@login_required(login_url="/login")
def detailed_course(request):
    context = {"breadcrumb": {"parent": "Learning", "child": "Detailed Course"}}
    return render(request, 'learning/detailed_course/learning-detailed.html', context)


# maps views


@login_required(login_url="/login")
def maps_js(request):
    context = {"breadcrumb": {"parent": "Maps", "child": "Map JS"}}
    return render(request, 'maps/maps_js/map-js.html', context)


@login_required(login_url="/login")
def vector_maps(request):
    context = {"breadcrumb": {"parent": "Maps", "child": "Vector Maps"}}
    return render(request, 'maps/vector_maps/vector-map.html', context)


# editors views


@login_required(login_url="/login")
def summer_note(request):
    context = {"breadcrumb": {"parent": "Editors", "child": "Summer Note"}}
    return render(request, 'editors/summer_note/summernote.html', context)


@login_required(login_url="/login")
def ck_editor(request):
    context = {"breadcrumb": {"parent": "Editors", "child": "Ck Editor"}}
    return render(request, 'editors/ck_editor/ckeditor.html', context)


@login_required(login_url="/login")
def mde_editor(request):
    context = {"breadcrumb": {"parent": "Editors", "child": "MDE Editor"}}
    return render(request, 'editors/mde_editor/simple-MDE.html', context)


@login_required(login_url="/login")
def ace_code_editor(request):
    context = {"breadcrumb": {"parent": "Editors", "child": "ACE Code Editor"}}
    return render(request, 'editors/ace_code_editor/ace-code-editor.html', context)


# knowledgebase views


@login_required(login_url="/login")
def knowledgebase(request):
    context = {"breadcrumb": {"parent": "Knowledgebase", "child": "Knowledgebase"}}
    return render(request, 'knowledgebase/knowledgebase/knowledgebase.html', context)


@login_required(login_url="/login")
def knowledge_category(request):
    context = {"breadcrumb": {"parent": "Knowledgebase", "child": "Knowledge Catagory"}}
    return render(request, 'knowledgebase/knowledge_category/knowledge-category.html', context)


@login_required(login_url="/login")
def knowledge_detail(request):
    context = {"breadcrumb": {"parent": "Knowledgebase", "child": "Knowledge Detail"}}
    return render(request, 'knowledgebase/knowledge_detail/knowledge-detail.html', context)


# support ticket


@login_required(login_url="/login")
def support_ticket(request):
    context = {"breadcrumb": {"parent": "Apps", "child": "Support Ticket"}}
    return render(request, 'support_ticket/support-ticket.html', context)


# documentation


@login_required(login_url="/login")
def documentation_accordian(request):
    return render(request, 'documentation/accordian.html')


@login_required(login_url="/login")
def documentation_app(request):
    return render(request, 'documentation/app.html')


@login_required(login_url="/login")
def documentation_change_log(request):
    return render(request, 'documentation/change-log.html')


@login_required(login_url="/login")
def documentation_component(request):
    return render(request, 'documentation/component.html')


@login_required(login_url="/login")
def documentation_customer_review(request):
    return render(request, 'documentation/customer-review.html')


@login_required(login_url="/login")
def documentation_feature_list(request):
    return render(request, 'documentation/feature-list.html')


@login_required(login_url="/login")
def documentation_getting_started(request):
    return render(request, 'documentation/getting-started.html')


@login_required(login_url="/login")
def documentation_index(request):
    return render(request, 'documentation/index.html')


@login_required(login_url="/login")
def documentation_layout_setting(request):
    return render(request, 'documentation/layout-setting.html')


@login_required(login_url="/login")
def documentation_options(request):
    return render(request, 'documentation/options.html')


@login_required(login_url="/login")
def documentation_tree(request):
    return render(request, 'documentation/tree.html')


@login_required(login_url="/login")
def documentation_django_to_do(request):
    return render(request, 'documentation/django-to-do.html')


@login_required(login_url="/login")
def documentation_django_authentication(request):
    return render(request, 'documentation/django-authentication.html')


@login_required(login_url="/login")
def documentation_django_customizer(request):
    return render(request, 'documentation/django-customizer.html')


@login_required(login_url="/login")
def documentation_django_components(request):
    return render(request, 'documentation/django-component.html')


@login_required(login_url="/login")
def documentation_django_getting_started(request):
    return render(request, 'documentation/django-getting-started.html')


@login_required(login_url="/login")
def documentation_django_options(request):
    return render(request, 'documentation/django-options.html')


@login_required(login_url="/login")
def documentation_django_tree(request):
    return render(request, 'documentation/django-tree.html')


@login_required(login_url="/login")
def documentation_django_app(request):
    return render(request, 'documentation/django-app.html')


@login_required(login_url="/login")
def documentation_django_app(request):
    return render(request, 'documentation/django-app.html')


# Carreras.
@login_required(login_url="/login")
def carreras_listas(request):
    lista_carreras = Carreras.objects.all()
    print(lista_carreras)
    context = {
        "breadcrumb": {"parent": "Carreras", "child": "Ver"},
        "lista_carreras": lista_carreras
    }
    return render(request, 'carreras/carreras/carreras.html', context)


@login_required(login_url="/login")
def carreras_agregar(request):
    if request.POST:
        form = CarreraForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('carreras')
    context = {
        "breadcrumb": {"parent": "Carreras", "child": "Añadir"},
        'form': CarreraForm
    }
    return render(request, 'carreras/carreras_agregar/carreras_agregar.html', context)


@login_required(login_url="/login")
def carreras_eliminar(request, pk):
    carrera = get_object_or_404(Carreras, idCarrera=pk)
    if request.POST:
        carrera.delete()
        return redirect('carreras')
    context = {
        "carrera": carrera,
        "breadcrumb": {"parent": "Carreras", "child": "Eliminar"}
    }
    return render(request, 'carreras/carreras_eliminar/carreras_eliminar.html', context)


@login_required(login_url="/login")
def carreras_modificar(request, pk):
    carrera = get_object_or_404(Carreras, idCarrera=pk)
    form = CarreraForm(initial={"nombreCarrera": carrera.nombreCarrera, "divisionCarrera": carrera.divisionCarrera})
    if request.POST:
        nomcarr = request.POST['nombreCarrera']
        divcarr = request.POST['divisionCarrera']
        carrera.nombreCarrera = nomcarr
        carrera.divisionCarrera = divcarr
        carrera.save()
        return redirect('carreras')
    context = {
        "breadcrumb": {"parent": "Carreras", "child": "Modificar"},
        'form': form,
        "Carrera": carrera,
    }
    return render(request, 'carreras/carreras_modificar/carreras_modificar.html', context)


@login_required(login_url="/login")
def carreras_ver(request, pk):
    carrera = get_object_or_404(Carreras, idCarrera=pk)
    context = {
        "breadcrumb": {"parent": "Carreras", "child": "Ver detalles"},
        "carrera": carrera,
    }
    return render(request, 'carreras/carreras_ver/carreras_ver.html', context)


# Empresas.
@login_required(login_url="/login")
def empresas_listas(request):
    lista_empresas = Empresa.objects.all()
    context = {
        "breadcrumb": {"parent": "Empresas", "child": "Ver"},
        "lista_empresas": lista_empresas
    }
    return render(request, 'empresas/empresas/empresas.html', context)


@login_required(login_url="/login")
def empresas_agregar(request):
    if request.POST:
        form = EmpresaForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('empresas')
    context = {
        "breadcrumb": {"parent": "Empresas", "child": "Añadir"},
        'form': EmpresaForm
    }
    return render(request, 'empresas/empresas_agregar/empresas_agregar.html', context)


@login_required(login_url="/login")
def empresas_modificar(request, pk):
    empresa = get_object_or_404(Empresa, idEmpresa=pk)
    form = EmpresaForm(initial={"razonSocial": empresa.razonSocial, "nombre": empresa.nombre,
                                "rfc": empresa.rfc, "giro": empresa.giro, "pais": empresa.pais,
                                "estado": empresa.estado,
                                "ciudad": empresa.ciudad, "colonia": empresa.colonia, "calle": empresa.calle,
                                "numero": empresa.numero, "numeroInterior": empresa.numeroInterior, "cp": empresa.cp,
                                "sectorEmpresa": empresa.sectorEmpresa})
    if request.POST:
        razonSocialEmpresa = request.POST['razonSocial']
        nomEmpresa = request.POST['nombre']
        rfcEmpresa = request.POST['rfc']
        giroEmpresa = request.POST['giro']
        pais = request.POST['pais']
        estado = request.POST['estado']
        ciudad = request.POST['ciudad']
        colonia = request.POST['colonia']
        calle = request.POST['calle']
        numero = request.POST['numero']
        numeroInterior = request.POST['numeroInterior']
        cp = request.POST['cp']
        sector = request.POST['sectorEmpresa']
        empresa.razonSocial = razonSocialEmpresa
        empresa.nombre = nomEmpresa
        empresa.rfc = rfcEmpresa
        empresa.giro = giroEmpresa
        empresa.pais = pais
        empresa.estado = estado
        empresa.ciudad = ciudad
        empresa.colonia = colonia
        empresa.calle = calle
        empresa.numero = numero
        empresa.numeroInterior = numeroInterior
        empresa.cp = cp
        empresa.sectorEmpresa = sector
        empresa.save()
        return redirect('empresas')
    context = {
        "breadcrumb": {"parent": "Empresas", "child": "Modificar"},
        'form': form,
        "Empresa": empresa,
    }
    return render(request, 'empresas/empresas_modificar/empresas_modificar.html', context)


@login_required(login_url="/login")
def empresas_eliminar(request, pk):
    empresa = get_object_or_404(Empresa, idEmpresa=pk)
    if request.POST:
        empresa.delete()
        return redirect('empresas')
    context = {
        "empresa": empresa,
        "breadcrumb": {"parent": "Empresas", "child": "Eliminar"}
    }
    return render(request, 'empresas/empresas_eliminar/empresas_eliminar.html', context)


@login_required(login_url="/login")
def empresas_ver(request, pk):
    empresa = get_object_or_404(Empresa, idEmpresa=pk)
    context = {
        "breadcrumb": {"parent": "Empresas", "child": "Ver detalles"},
        "empresa": empresa,
    }
    return render(request, 'empresas/empresas_ver/empresas_ver.html', context)


# Contactos.
# Contactos listas.
@login_required(login_url="/login")
def contactos_listas(request):
    lista_contactos = Contacto.objects.all()
    context = {
        "breadcrumb": {"parent": "Contactos", "child": "Ver"},
        "lista_contactos": lista_contactos
    }
    return render(request, 'contactos/contactos/contactos.html', context)


# Contactos agregar.
@login_required(login_url="/login")
def contactos_agregar(request):
    if request.POST:
        form = ContactoForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('contactos')
    context = {
        "breadcrumb": {"parent": "Contactos", "child": "Añadir"},
        'form': ContactoForm
    }
    return render(request, 'contactos/contactos_agregar/contactos_agregar.html', context)


# Contactos modificar.
@login_required(login_url="/login")
def contactos_modificar(request, pk):
    contacto = get_object_or_404(Contacto, idContacto=pk)
    form = ContactoForm(initial={"nombre": contacto.nombre, "numTelefono": contacto.numTelefono,
                                 "idEmpresa": contacto.idEmpresa.razonSocial, "email": contacto.email})

    if request.POST:
        nombre = request.POST['nombre']
        numTelefono = request.POST['numTelefono']
        empresa_cadena = request.POST['idEmpresa']
        email = request.POST['email']
        idEmpresa = Empresa.objects.get(idEmpresa=empresa_cadena)
        contacto.nombre = nombre
        contacto.numTelefono = numTelefono
        contacto.idEmpresa = idEmpresa
        contacto.email = email
        contacto.save()
        return redirect('contactos')
    context = {
        "breadcrumb": {"parent": "Contactos", "child": "Modificar"},
        'form': form,
        "Contacto": contacto,
    }
    return render(request, 'contactos/contactos_modificar/contactos_modificar.html', context)


# Contactos eliminar.
@login_required(login_url="/login")
def contactos_eliminar(request, pk):
    contacto = get_object_or_404(Contacto, idContacto=pk)
    if request.POST:
        contacto.delete()
        return redirect('contactos')
    context = {
        "contacto": contacto,
        "breadcrumb": {"parent": "Contactos", "child": "Eliminar"}
    }
    return render(request, 'contactos/contactos_eliminar/contactos_eliminar.html', context)


# Contactos ver.
@login_required(login_url="/login")
def contactos_ver(request, pk):
    contacto = get_object_or_404(Contacto, idContacto=pk)
    context = {
        "breadcrumb": {"parent": "Contactos", "child": "Ver detalles"},
        "contacto": contacto,
    }
    return render(request, 'contactos/contactos_ver/contactos_ver.html', context)


# Convenios.
# Convenios listas
@login_required(login_url="/login")
def convenios_listas(request):
    lista_convenios = Convenio.objects.all()
    context = {
        "breadcrumb": {"parent": "Convenios", "child": "Ver"},
        "lista_convenios": lista_convenios
    }
    return render(request, 'convenios/convenios/convenios.html', context)


# Convenios agregar
@login_required(login_url="/login")
def convenios_agregar(request):
    if request.POST:
        form = ConvenioForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('convenios')
    context = {
        "breadcrumb": {"parent": "Convenios", "child": "Añadir"},
        'form': ConvenioForm
    }
    return render(request, 'convenios/convenios_agregar/convenios_agregar.html', context)


# Convenios modificar
@login_required(login_url="/login")
def convenios_modificar(request, pk):
    convenio = get_object_or_404(Convenio, numConvenio=pk)
    form = ConvenioForm(initial={"idCarrera": convenio.idCarrera.nombreCarrera,
                                 "inicioVigencia": convenio.inicioVigencia,
                                 "idEmpresa": convenio.idEmpresa.razonSocial,
                                 "finVigencia": convenio.finVigencia,
                                 "observaciones": convenio.observaciones})
    if request.POST:
        carrera_cadena = request.POST['idCarrera']
        idCarrera = Carreras.objects.get(idCarrera=carrera_cadena)
        inicioVigencia = request.POST['inicioVigencia']
        finVigencia = request.POST['finVigencia']
        empresa_cadena = request.POST['idEmpresa']
        idEmpresa = Empresa.objects.get(idEmpresa=empresa_cadena)
        observaciones = request.POST['observaciones']
        convenio.idCarrera = idCarrera
        convenio.inicioVigencia = inicioVigencia
        convenio.finVigencia = finVigencia
        convenio.idEmpresa = idEmpresa
        convenio.observaciones = observaciones
        convenio.save()
        return redirect('convenios')
    context = {
        "breadcrumb": {"parent": "Convenios", "child": "Modificar"},
        'form': form,
        "Convenio": convenio,
    }
    return render(request, 'convenios/convenios_modificar/convenios_modificar.html', context)


# Convenios Eliminar
@login_required(login_url="/login")
def convenios_eliminar(request, pk):
    convenio = get_object_or_404(Convenio, numConvenio=pk)
    if request.POST:
        convenio.delete()
        return redirect('convenios')
    context = {
        "convenio": convenio,
        "breadcrumb": {"parent": "Convenio", "child": "Eliminar"}
    }
    return render(request, 'convenios/convenios_eliminar/convenios_eliminar.html', context)


# Convenios Ver
@login_required(login_url="/login")
def convenios_ver(request, pk):
    convenio = get_object_or_404(Convenio, numConvenio=pk)
    context = {
        "breadcrumb": {"parent": "Contactos", "child": "Ver detalles"},
        "convenio": convenio,
    }
    return render(request, 'convenios/convenios_ver/convenios_ver.html', context)


# Módulo de reportes

from functools import reduce
from operator import or_
from django.db.models import Q
from django.db.models import QuerySet


# Módulo de reportes
# views.py
def reportes(request):
    lista_convenios = Convenio.objects.all().order_by('-inicioVigencia')
    empresas = Empresa.objects.filter(convenio__isnull=False).distinct()
    carreras = Carreras.objects.filter(convenio__isnull=False).distinct()
    error_message = None
    earliest_convenio = Convenio.objects.order_by('inicioVigencia').first()
    min_date = earliest_convenio.inicioVigencia if earliest_convenio else None

    form = ReporteConveniosForm(request.GET, initial={'fecha_inicio': min_date})

    if request.method == "GET" and form.is_valid():
        cleaned_data = form.cleaned_data
        idCarrera = cleaned_data['idCarrera']
        idEmpresa = cleaned_data['idEmpresa']
        fecha_inicio = cleaned_data['fecha_inicio']
        fecha_vigencia = cleaned_data['fecha_vigencia']
        estado = cleaned_data['estado']
        print("Fecha de vigencia:", fecha_vigencia)

        q_objects = Q()

        if idCarrera:
            q_objects &= reduce(or_, [Q(idCarrera=carrera) for carrera in idCarrera])

        if idEmpresa:
            q_objects &= reduce(or_, [Q(idEmpresa=empresa) for empresa in idEmpresa])

        if fecha_inicio:
            q_objects &= Q(inicioVigencia__gte=fecha_inicio)

        if fecha_vigencia is not None and fecha_vigencia != 'Todas las fechas':
            rango_fechas = fecha_vigencia.split(' - ')
            if len(rango_fechas) == 2:
                fecha_inicio_vigencia = datetime.strptime(rango_fechas[0], '%Y-%m-%d')
                fecha_fin_vigencia = datetime.strptime(rango_fechas[1], '%Y-%m-%d')
                q_objects &= (Q(inicioVigencia__range=(fecha_inicio_vigencia, fecha_fin_vigencia)) |
                              Q(finVigencia__range=(fecha_inicio_vigencia, fecha_fin_vigencia)))
            else:
                error_message = "El formato de fecha debe ser 'DD/MM/YYYY - DD/MM/YYYY'."

        if estado == 'activo':
            q_objects &= Q(inicioVigencia__lte=timezone.now().date(), finVigencia__gte=timezone.now().date())
        elif estado == 'casi_expirado':
            q_objects &= Q(finVigencia__gt=timezone.now().date(),
                           finVigencia__lte=timezone.now().date() + timedelta(days=30))
        elif estado == 'expirado':
            q_objects &= Q(finVigencia__lt=timezone.now().date())
        lista_convenios = lista_convenios.filter(q_objects)

    context = {
        'form': form,
        'lista_convenios': lista_convenios,
        'empresas': empresas,
        'carreras': carreras,
        'error_message': error_message,
        'min_date': min_date,
        'breadcrumb': {"parent": "Reportes", "child": "Reportes de convenios"}
    }
    return render(request, 'dashboard/main_dashboard.html', context)


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime

def export_convenios_excel(request):
    convenios = Convenio.objects.all().order_by('-inicioVigencia')

    # Crear el archivo de Excel y agregar una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Convenios"

    # Agregar encabezados
    headers = [
        'ID Convenio', 'Estado',
        'Carrera', 'ID Carrera', 'División Carrera', 'Empresa',
        'Nombre Empresa', 'ID Empresa', 'Giro Empresa', 'Sector Empresa',
        'Inicio de vigencia', 'Fin de vigencia', 'Observaciones'
    ]
    # Definir los estilos de celda para los estados de los convenios
    activo_fill = PatternFill(start_color="c3e6cb", end_color="c3e6cb", fill_type="solid")
    casi_expirado_fill = PatternFill(start_color="ffeeba", end_color="ffeeba", fill_type="solid")
    expirado_fill = PatternFill(start_color="f5c6cb", end_color="f5c6cb", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet['{}1'.format(col_letter)] = header
        worksheet.column_dimensions[col_letter].width = 20

    # Llenar la hoja de trabajo con los datos de convenios
    for row_num, convenio in enumerate(convenios, 2):
        estado = convenio.estado
        print(estado)
        row_data = [
            convenio.numConvenio,
            estado,
            convenio.idCarrera.nombreCarrera,
            convenio.idCarrera.idCarrera,
            convenio.idCarrera.divisionCarrera,
            convenio.idEmpresa.razonSocial,
            convenio.idEmpresa.nombre,
            convenio.idEmpresa.idEmpresa,
            convenio.idEmpresa.giro,
            convenio.idEmpresa.sectorEmpresa,
            convenio.inicioVigencia,
            convenio.finVigencia,
            convenio.observaciones,
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

            # Aplicar el estilo según el estado del convenio
            if estado == "Activo":
                cell.fill = activo_fill
            elif estado == "Casi expirado":
                cell.fill = casi_expirado_fill
            else:
                cell.fill = expirado_fill

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Formatear la fecha y hora actual como una cadena y usarla en el nombre del archivo
    now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f'convenios_{now}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    workbook.save(response)

    return response

# Dashboard
from django.shortcuts import render
from datetime import datetime, timedelta
from .models import Convenio, Empresa, Carreras
from django.db.models import Count
from django.db.models.functions import TruncMonth
from datetime import timedelta

def dashboard(request):
    today = datetime.now().date()
    one_month_ago = today - timedelta(days=30)
    three_months_ago = today - timedelta(days=90)
    six_months_ago = today - timedelta(days=180)
    one_year_ago = today - timedelta(days=365)

    # Convenios por estado
    convenios_activos = Convenio.objects.filter(finVigencia__gte=today).count()
    convenios_casi_expirados = Convenio.objects.filter(finVigencia__lt=today, finVigencia__gte=one_month_ago).count()
    convenios_expirados = Convenio.objects.filter(finVigencia__lt=one_month_ago).count()

    # Convenios por fecha de inicio
    convenios_30_dias = Convenio.objects.filter(inicioVigencia__gte=one_month_ago).count()
    convenios_3_meses = Convenio.objects.filter(inicioVigencia__gte=three_months_ago).count()
    convenios_6_meses = Convenio.objects.filter(inicioVigencia__gte=six_months_ago).count()
    convenios_1_ano = Convenio.objects.filter(inicioVigencia__gte=one_year_ago).count()

    # Convenios por carrera
    convenios_por_carrera = Carreras.objects.all().annotate(convenios_count=Count('convenio')).order_by(
        '-convenios_count')[:5].values('nombreCarrera', 'convenios_count')


    # Convenios próximos a expirar
    convenios_proximos_expirar = Convenio.objects.filter(finVigencia__lte=one_month_ago).order_by('finVigencia')

    now = timezone.now()
    last_year = now - timedelta(days=365)

    convenios_por_mes = (
        Convenio.objects.filter(inicioVigencia__gte=last_year)
        .annotate(month=TruncMonth("inicioVigencia"))
        .values("month")
        .annotate(convenios_count=Count("numConvenio"))
        .order_by("month")
    )
    #Numero total de convenios
    numero_total_convenios = Convenio.objects.all().count()

    #Numero de empresas con convenios activos
    empresas_convenios_activos = Empresa.objects.filter(convenio__finVigencia__gte=today).distinct().count()


    # ... Realiza otras consultas y cálculos necesarios ...

    context = {
        'convenios_activos': convenios_activos,
        'convenios_casi_expirados': convenios_casi_expirados,
        'convenios_expirados': convenios_expirados,
        'convenios_30_dias': convenios_30_dias,
        'convenios_3_meses': convenios_3_meses,
        'convenios_6_meses': convenios_6_meses,
        'convenios_1_ano': convenios_1_ano,
        'convenios_por_carrera': convenios_por_carrera,
        'convenios_proximos_expirar': convenios_proximos_expirar,
        'convenios_por_carrera_json': json.dumps(list(convenios_por_carrera)),
        'convenios_por_mes_json': json.dumps(list(convenios_por_mes), cls=DjangoJSONEncoder),
        'numero_total_convenios': numero_total_convenios,
        'empresas_convenios_activos': empresas_convenios_activos,
        'breadcrumb': {"parent": "Tablero de inicio", "child": "Tablero de inicio"},
        # ... Pasa otras variables de contexto necesarias ...
    }
    print(json.dumps(list(convenios_por_carrera)))
    return render(request, 'dashboard/main_dashboard.html', context)
